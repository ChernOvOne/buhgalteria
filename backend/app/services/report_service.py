from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics.shapes import Drawing, Rect
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import date, datetime
from typing import List, Dict, Any, Optional
from babel.dates import format_date


# ── PDF Report ────────────────────────────────────────────────────────────────

def generate_pdf_report(
    company_name: str,
    period_label: str,
    kpi: Dict,
    transactions: List[Dict],
    expense_by_category: List[Dict],
    partners_summary: List[Dict],
    ad_stats: Optional[Dict] = None,
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    PRIMARY = colors.HexColor("#534AB7")
    SUCCESS = colors.HexColor("#1D9E75")
    DANGER = colors.HexColor("#E24B4A")
    GRAY = colors.HexColor("#888780")
    LIGHT = colors.HexColor("#F1EFE8")

    title_style = ParagraphStyle("title", parent=styles["Normal"],
                                  fontSize=20, textColor=PRIMARY, spaceAfter=4,
                                  fontName="Helvetica-Bold")
    h2_style = ParagraphStyle("h2", parent=styles["Normal"],
                               fontSize=13, textColor=PRIMARY, spaceBefore=16, spaceAfter=8,
                               fontName="Helvetica-Bold")
    normal = ParagraphStyle("n", parent=styles["Normal"], fontSize=10, textColor=colors.black)
    muted = ParagraphStyle("m", parent=styles["Normal"], fontSize=9, textColor=GRAY)

    story = []

    # Header
    story.append(Paragraph(company_name, title_style))
    story.append(Paragraph(f"Финансовый отчёт · {period_label}", muted))
    story.append(Paragraph(f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}", muted))
    story.append(HRFlowable(width="100%", thickness=1, color=PRIMARY, spaceAfter=12))

    # KPI Summary
    story.append(Paragraph("Ключевые показатели", h2_style))

    def fmt(v):
        return f"{v:,.2f} ₽".replace(",", " ")

    kpi_data = [
        ["Показатель", "Значение"],
        ["Выручка (доход)", fmt(kpi.get("income", 0))],
        ["Расходы", fmt(kpi.get("expense", 0))],
        ["Чистая прибыль", fmt(kpi.get("profit", 0))],
        ["Средняя выручка в день", fmt(kpi.get("avg_per_day", 0))],
    ]
    if kpi.get("best_day"):
        kpi_data.append(["Лучший день", f"{kpi['best_day']} — {fmt(kpi.get('best_day_amount', 0))}"])

    kpi_table = Table(kpi_data, colWidths=[9*cm, 8*cm])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D3D1C7")),
        ("BACKGROUND", (0, 2), (-1, 2), LIGHT),
        ("BACKGROUND", (0, 4), (-1, 4), LIGHT),
        ("TEXTCOLOR", (1, 3), (1, 3), SUCCESS if kpi.get("profit", 0) >= 0 else DANGER),
        ("FONTNAME", (1, 3), (1, 3), "Helvetica-Bold"),
        ("PADDING", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
    ]))
    story.append(kpi_table)

    # Расходы по категориям
    if expense_by_category:
        story.append(Paragraph("Расходы по категориям", h2_style))
        cat_data = [["Категория", "Сумма", "% от расходов"]]
        total_exp = sum(c.get("amount", 0) for c in expense_by_category)
        for c in expense_by_category:
            pct = round(c["amount"] / total_exp * 100, 1) if total_exp > 0 else 0
            cat_data.append([c["name"], fmt(c["amount"]), f"{pct}%"])
        cat_data.append(["Итого", fmt(total_exp), "100%"])

        cat_table = Table(cat_data, colWidths=[9*cm, 5*cm, 3*cm])
        cat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), LIGHT),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D3D1C7")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, LIGHT]),
            ("PADDING", (0, 0), (-1, -1), 8),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ]))
        story.append(cat_table)

    # Партнёры
    if partners_summary:
        story.append(Paragraph("Партнёры и дивиденды", h2_style))
        p_data = [["Участник", "Роль", "Последние ДВД", "Остаток долга"]]
        for p in partners_summary:
            p_data.append([
                p.get("name", ""),
                p.get("role_label", ""),
                fmt(p.get("last_dividend") or 0) if p.get("last_dividend") else "—",
                fmt(p.get("remaining_debt") or 0),
            ])
        p_table = Table(p_data, colWidths=[5*cm, 4*cm, 4.5*cm, 3.5*cm])
        p_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D3D1C7")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
            ("PADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(p_table)

    # Транзакции
    if transactions:
        story.append(Paragraph("Детализация транзакций", h2_style))
        t_data = [["Дата", "Тип", "Категория", "Описание", "Сумма"]]
        for t in transactions[:200]:  # максимум 200 строк в PDF
            t_type = "Доход" if t.get("type") == "income" else "Расход"
            cat_name = t.get("category", {}).get("name", "—") if t.get("category") else "—"
            t_data.append([
                str(t.get("date", "")),
                t_type,
                cat_name,
                (t.get("description") or "")[:40],
                fmt(t.get("amount", 0)),
            ])

        t_table = Table(t_data, colWidths=[2.5*cm, 2*cm, 3*cm, 7*cm, 2.5*cm])
        t_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D3D1C7")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
            ("PADDING", (0, 0), (-1, -1), 5),
            ("ALIGN", (4, 0), (4, -1), "RIGHT"),
        ]))
        story.append(t_table)

    # Footer
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=0.5, color=GRAY))
    story.append(Paragraph(
        f"Отчёт сформирован автоматически системой бухгалтерии · {company_name}",
        ParagraphStyle("footer", parent=muted, alignment=1)
    ))

    doc.build(story)
    return buf.getvalue()


# ── Excel Report ──────────────────────────────────────────────────────────────

def generate_excel_report(
    company_name: str,
    date_from: date,
    date_to: date,
    transactions: List[Dict],
    expense_by_category: List[Dict],
    inkas_records: List[Dict],
    ad_campaigns: List[Dict],
) -> bytes:
    wb = Workbook()

    # Стили
    header_fill = PatternFill("solid", fgColor="534AB7")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    income_fill = PatternFill("solid", fgColor="EAF3DE")
    expense_fill = PatternFill("solid", fgColor="FCEBEB")
    alt_fill = PatternFill("solid", fgColor="F1EFE8")
    total_fill = PatternFill("solid", fgColor="D3D1C7")
    total_font = Font(bold=True, size=11)
    thin = Border(
        left=Side(style="thin", color="D3D1C7"),
        right=Side(style="thin", color="D3D1C7"),
        top=Side(style="thin", color="D3D1C7"),
        bottom=Side(style="thin", color="D3D1C7"),
    )
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right")

    def set_header(ws, headers, col_widths=None):
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin
        if col_widths:
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

    def apply_border(ws, max_row, max_col):
        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = thin

    # ── Лист 1: Транзакции ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Транзакции"
    ws1.freeze_panes = "A2"

    set_header(ws1,
        ["Дата", "Тип", "Категория", "Описание", "Сумма", "Чек"],
        [14, 10, 16, 40, 14, 50]
    )

    row = 2
    for t in transactions:
        is_income = t.get("type") == "income"
        cells_data = [
            t.get("date", ""),
            "Доход" if is_income else "Расход",
            (t.get("category") or {}).get("name", ""),
            t.get("description", "") or "",
            t.get("amount", 0),
            t.get("receipt_url", "") or "",
        ]
        for col, val in enumerate(cells_data, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.fill = income_fill if is_income else expense_fill
            if col == 5:
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
        row += 1

    # Итого
    ws1.cell(row=row, column=1, value="ИТОГО").font = total_font
    income_total = sum(t.get("amount", 0) for t in transactions if t.get("type") == "income")
    expense_total = sum(t.get("amount", 0) for t in transactions if t.get("type") == "expense")
    ws1.cell(row=row, column=4, value="Доход:").font = total_font
    ws1.cell(row=row, column=5, value=income_total).font = total_font
    ws1.cell(row=row+1, column=4, value="Расход:").font = total_font
    ws1.cell(row=row+1, column=5, value=expense_total).font = total_font
    ws1.cell(row=row+2, column=4, value="Прибыль:").font = total_font
    profit_cell = ws1.cell(row=row+2, column=5, value=income_total - expense_total)
    profit_cell.font = total_font

    apply_border(ws1, row - 1, 6)

    # ── Лист 2: Расходы по категориям ─────────────────────────────────────
    ws2 = wb.create_sheet("Расходы по категориям")
    set_header(ws2, ["Категория", "Сумма", "% от расходов"], [20, 16, 16])

    total_exp = sum(c.get("amount", 0) for c in expense_by_category)
    for i, c in enumerate(expense_by_category, 2):
        pct = round(c["amount"] / total_exp * 100, 1) if total_exp > 0 else 0
        ws2.cell(row=i, column=1, value=c["name"])
        amt_cell = ws2.cell(row=i, column=2, value=c["amount"])
        amt_cell.number_format = '#,##0.00'
        ws2.cell(row=i, column=3, value=pct / 100).number_format = "0.0%"
        if i % 2 == 0:
            for col in range(1, 4):
                ws2.cell(row=i, column=col).fill = alt_fill

    r = len(expense_by_category) + 2
    ws2.cell(row=r, column=1, value="ИТОГО").fill = total_fill
    ws2.cell(row=r, column=1).font = total_font
    ws2.cell(row=r, column=2, value=total_exp).fill = total_fill
    ws2.cell(row=r, column=2).font = total_font
    ws2.cell(row=r, column=2).number_format = '#,##0.00'

    apply_border(ws2, r, 3)

    # ── Лист 3: Инкас ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Инкас")
    set_header(ws3, ["Дата", "Период", "Тип", "Партнёр", "Сумма"], [14, 14, 16, 16, 14])

    type_labels = {"dividend": "ДВД", "return_inv": "ВОЗВРИНВ", "investment": "ВЛОЖЕНИЕ"}
    for i, r in enumerate(inkas_records, 2):
        ws3.cell(row=i, column=1, value=str(r.get("date", "")))
        ws3.cell(row=i, column=2, value=r.get("month_label", ""))
        ws3.cell(row=i, column=3, value=type_labels.get(r.get("type", ""), r.get("type", "")))
        ws3.cell(row=i, column=4, value=r.get("partner_name", ""))
        amt = ws3.cell(row=i, column=5, value=r.get("amount", 0))
        amt.number_format = '#,##0.00'
        if i % 2 == 0:
            for col in range(1, 6):
                ws3.cell(row=i, column=col).fill = alt_fill

    apply_border(ws3, len(inkas_records) + 1, 5)

    # ── Лист 4: Реклама ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("Реклама")
    set_header(ws4,
        ["Дата", "Канал", "Формат", "Сумма", "ПДП", "₽/ПДП", "Ссылка"],
        [14, 30, 12, 14, 10, 10, 50]
    )

    for i, a in enumerate(ad_campaigns, 2):
        ws4.cell(row=i, column=1, value=str(a.get("date", "")))
        ws4.cell(row=i, column=2, value=a.get("channel_name", "") or "")
        ws4.cell(row=i, column=3, value=a.get("format", "") or "")
        amt = ws4.cell(row=i, column=4, value=a.get("amount", 0))
        amt.number_format = '#,##0.00'
        ws4.cell(row=i, column=5, value=a.get("subscribers_gained") or "")
        cps = a.get("cost_per_sub")
        ws4.cell(row=i, column=6, value=round(cps, 2) if cps else "")
        ws4.cell(row=i, column=7, value=a.get("channel_url", "") or "")
        if i % 2 == 0:
            for col in range(1, 8):
                ws4.cell(row=i, column=col).fill = alt_fill

    apply_border(ws4, len(ad_campaigns) + 1, 7)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
